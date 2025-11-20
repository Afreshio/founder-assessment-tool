"""
Generate ScaleOS Financial Model Excel file with corrected Active_CEOs calculation.
Uses SUMPRODUCT instead of OFFSET to prevent flatlining.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

def create_scaleos_model():
    """Create the ScaleOS Financial Model Excel workbook."""
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"
    
    # Define styles
    header_font = Font(bold=True, size=11)
    input_font = Font(size=10)
    number_format = '#,##0'
    currency_format = '$#,##0'
    
    # Inputs section (rows 1-5)
    ws['E1'] = "Inputs"
    ws['E1'].font = Font(bold=True, size=12)
    
    inputs = [
        ("Tenure (months)", 6),
        ("Initial CEOs (Month 1)", 2),
        ("New CEOs added / month", 1),
        ("On-Demand price per course", 500),
    ]
    
    for idx, (label, default_value) in enumerate(inputs, start=2):
        ws[f'E{idx}'] = label
        ws[f'E{idx}'].font = input_font
        ws[f'F{idx}'] = default_value
        ws[f'F{idx}'].font = input_font
        ws[f'F{idx}'].number_format = number_format
    
    # Summary block (rows 7-12)
    ws['D7'] = "CEO Revenue ($)"
    ws['D8'] = "Workshop Revenue ($)"
    ws['D9'] = "Sprint Revenue ($)"
    ws['D10'] = "On-Demand Revenue ($)"
    ws['D11'] = "Total Revenue ($)"
    ws['D12'] = "Days"
    
    for row in range(7, 13):
        ws[f'D{row}'].font = header_font
    
    # Year headers for summary
    ws['E6'] = 2026
    ws['F6'] = 2027
    ws['G6'] = 2028
    for col in ['E', 'F', 'G']:
        ws[f'{col}6'].font = header_font
        ws[f'{col}6'].alignment = Alignment(horizontal='center')
    
    # Column headers (row 14)
    headers = [
        "Year",
        "Month",
        "New_CEOs",
        "Active_CEOs",
        "CEO Revenue ($)",
        "Workshops",
        "Workshop Revenue ($)",
        "Sprints",
        "Sprint Revenue ($)",
        "CEO Hours",
        "Workshop Hours",
        "Sprint Hours",
        "Total Hours",
        "Total Days",
        "Total Revenue ($)",
        "On-Demand Courses",
        "On-Demand Revenue ($)",
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=14, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Generate 36 months of data (Jan 2026 - Dec 2028)
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    
    start_row = 15
    end_row = start_row + 36 - 1
    
    for i in range(36):
        row = start_row + i
        year = 2026 + (i // 12)
        month_idx = i % 12
        month_name = months[month_idx]
        
        # Year (Column A)
        ws[f'A{row}'] = year
        
        # Month (Column B)
        ws[f'B{row}'] = month_name
        
        # New_CEOs (Column C)
        if i == 0:
            # First month uses Initial CEOs
            ws[f'C{row}'] = f'=$F$3'
        else:
            # Other months use New CEOs per month
            ws[f'C{row}'] = f'=$F$4'
        
        # Active_CEOs (Column D) - CRITICAL FIX using SUMPRODUCT (no OFFSET)
        # This formula sums New_CEOs from the last 'Tenure' months
        # Formula structure: =SUMPRODUCT($C$15:C[row], --(ROW($C$15:C[row]) >= ROW(C[row])-$F$2+1))
        # This ensures we only count CEOs from the last tenure months
        # The -- converts TRUE/FALSE to 1/0, then SUMPRODUCT multiplies and sums
        first_data_row = start_row
        current_row_ref = f'C{row}'
        # Exact formula from spec: SUMPRODUCT(values, condition)
        ws[f'D{row}'] = f'=SUMPRODUCT($C${first_data_row}:{current_row_ref},--(ROW($C${first_data_row}:{current_row_ref})>=ROW({current_row_ref})-$F$2+1))'
        
        # CEO Revenue (Column E)
        ws[f'E{row}'] = f'=D{row}*5000'
        ws[f'E{row}'].number_format = currency_format
        
        # Workshops (Column F)
        # Base: 1 per month
        # Bonus: +1 in March, June, September, December (month indices 2, 5, 8, 11)
        if month_idx in [2, 5, 8, 11]:  # Mar, Jun, Sep, Dec
            ws[f'F{row}'] = 2
        else:
            ws[f'F{row}'] = 1
        
        # Workshop Revenue (Column G)
        ws[f'G{row}'] = f'=F{row}*7500'
        ws[f'G{row}'].number_format = currency_format
        
        # Sprints (Column H)
        # 1 sprint in January, April, July, October (month indices 0, 3, 6, 9)
        if month_idx in [0, 3, 6, 9]:  # Jan, Apr, Jul, Oct
            ws[f'H{row}'] = 1
        else:
            ws[f'H{row}'] = 0
        
        # Sprint Revenue (Column I)
        ws[f'I{row}'] = f'=H{row}*25000'
        ws[f'I{row}'].number_format = currency_format
        
        # CEO Hours (Column J)
        ws[f'J{row}'] = f'=D{row}*2*1.5'
        ws[f'J{row}'].number_format = '0.0'
        
        # Workshop Hours (Column K)
        ws[f'K{row}'] = f'=F{row}*4'
        ws[f'K{row}'].number_format = '0.0'
        
        # Sprint Hours (Column L)
        ws[f'L{row}'] = f'=H{row}*17'
        ws[f'L{row}'].number_format = '0.0'
        
        # Total Hours (Column M)
        ws[f'M{row}'] = f'=J{row}+K{row}+L{row}'
        ws[f'M{row}'].number_format = '0.0'
        
        # Total Days (Column N)
        ws[f'N{row}'] = f'=M{row}/8'
        ws[f'N{row}'].number_format = '0.0'
        
        # On-Demand Courses (Column P)
        ws[f'P{row}'] = 0
        
        # On-Demand Revenue (Column Q)
        ws[f'Q{row}'] = f'=P{row}*$F$5'
        ws[f'Q{row}'].number_format = currency_format
        
        # Total Revenue (Column O)
        ws[f'O{row}'] = f'=E{row}+G{row}+I{row}+Q{row}'
        ws[f'O{row}'].number_format = currency_format
    
    # Yearly summary formulas (2026, 2027, 2028)
    summary_rows = {
        'CEO Revenue': 7,
        'Workshop Revenue': 8,
        'Sprint Revenue': 9,
        'On-Demand Revenue': 10,
        'Total Revenue': 11,
        'Days': 12,
    }
    
    summary_columns = {
        'CEO Revenue': 'E',
        'Workshop Revenue': 'G',
        'Sprint Revenue': 'I',
        'On-Demand Revenue': 'Q',
        'Total Revenue': 'O',
        'Days': 'N',
    }
    
    for year_offset, year_col in enumerate(['E', 'F', 'G'], start=0):
        year = 2026 + year_offset
        
        # CEO Revenue
        ws[f'{year_col}7'] = f'=SUMIFS($E${start_row}:$E${end_row},$A${start_row}:$A${end_row},{year})'
        ws[f'{year_col}7'].number_format = currency_format
        
        # Workshop Revenue
        ws[f'{year_col}8'] = f'=SUMIFS($G${start_row}:$G${end_row},$A${start_row}:$A${end_row},{year})'
        ws[f'{year_col}8'].number_format = currency_format
        
        # Sprint Revenue
        ws[f'{year_col}9'] = f'=SUMIFS($I${start_row}:$I${end_row},$A${start_row}:$A${end_row},{year})'
        ws[f'{year_col}9'].number_format = currency_format
        
        # On-Demand Revenue
        ws[f'{year_col}10'] = f'=SUMIFS($Q${start_row}:$Q${end_row},$A${start_row}:$A${end_row},{year})'
        ws[f'{year_col}10'].number_format = currency_format
        
        # Total Revenue
        ws[f'{year_col}11'] = f'=SUMIFS($O${start_row}:$O${end_row},$A${start_row}:$A${end_row},{year})'
        ws[f'{year_col}11'].number_format = currency_format
        
        # Days
        ws[f'{year_col}12'] = f'=SUMIFS($N${start_row}:$N${end_row},$A${start_row}:$A${end_row},{year})'
        ws[f'{year_col}12'].number_format = '0.0'
    
    # Adjust column widths
    column_widths = {
        'A': 8,   # Year
        'B': 10,  # Month
        'C': 12,  # New_CEOs
        'D': 14,  # Active_CEOs
        'E': 16,  # CEO Revenue
        'F': 12,  # Workshops
        'G': 18,  # Workshop Revenue
        'H': 10,  # Sprints
        'I': 16,  # Sprint Revenue
        'J': 12,  # CEO Hours
        'K': 15,  # Workshop Hours
        'L': 12,  # Sprint Hours
        'M': 13,  # Total Hours
        'N': 12,  # Total Days
        'O': 16,  # Total Revenue
        'P': 18,  # On-Demand Courses
        'Q': 20,  # On-Demand Revenue
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Save the workbook
    filename = 'ScaleOS_Financial_Model.xlsx'
    wb.save(filename)
    print(f"✅ Excel model created: {filename}")
    print(f"   - Active_CEOs uses SUMPRODUCT formula (no OFFSET)")
    print(f"   - Model spans {start_row} to {end_row} (36 months)")
    print(f"   - Years: 2026, 2027, 2028")
    print(f"   - Ready to test - change inputs in F2-F5 to verify no flatlining")

if __name__ == '__main__':
    create_scaleos_model()

